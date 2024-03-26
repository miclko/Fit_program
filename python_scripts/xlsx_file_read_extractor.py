import openpyxl
from typing import List

def Get_workbook_from_xlsx_file(filename, be_silent = False):
    '''Get the workbook of the chosen xlsx file, print also which sheets it has and how to access them'''
    if not isinstance(filename,str):
        raise TypeError('no string as filename given')
    workbook = openpyxl.load_workbook(filename +".xlsx",read_only=True)
    sheetnames = workbook.sheetnames
    if not be_silent:
        print("open xlsx file " + filename)
        print("sheetnames:")
        print(sheetnames )
        print("returns wb. sheets can be accessed with wb[wb.sheetnames[i]], i in range(len(wb.sheetnames))")
    return workbook

def Get_data_from_sheet(sheet):
    '''Get all rows as data dict. Assumes certain structure as given by Neil
       Input: sheet, sheet being considered
       Output: Dict of all rows
    '''
    if not isinstance(sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        raise TypeError('works with an openpyxl ReadOnlyWorksheet')
    outputdict = Get_data_from_single_sheet_column(sheet,0)
    for x in range(1, sheet.max_column):
        outputdict.update(Get_data_from_single_sheet_column(sheet,x))
    return outputdict
    



def Get_data_from_single_sheet_column(sheet:openpyxl.worksheet._read_only.ReadOnlyWorksheet, columnNr:int) -> dict:
    '''Get a single Column dict from a given sheet
       Input: sheet, column number
       Output: dict with a single element'''
    columnCount = sheet.max_column
    rowCount = sheet.max_row
    if columnNr >= columnCount:
        raise ValueError('too large column chosen')
    columnsconverter = ['A','B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    cells = sheet[f'{columnsconverter[columnNr]}2':f'{columnsconverter[columnNr]}{rowCount}']
    valList =  []
    for x in cells[1:]:
        valList.append(x[0].value)
    return {cells[0][0].value : valList}


